"""Bulk dish Excel import + image filename mapping."""

import io
import zipfile

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook

from app.dish_bulk import BULK_HEADERS, build_dish_bulk_template_xlsx


def _minimal_jpeg() -> bytes:
    # 1x1 JPEG
    return bytes.fromhex(
        "ffd8ffe000104a46494600010100000100010000ffdb004300080606070605080707"
        "070909080a0c140d0c0b0b0c1912130f141d1a1f1e1d1a1c1c20242e2720222c231c"
        "1c2837292c30313434341f27393d38323c2e333432ffdb0043010909090c0b0c180d"
        "0d1832211c2132323232323232323232323232323232323232323232323232323232"
        "323232323232323232323232323232323232323232ffc00011080001000103011100"
        "021101031101ffc40014000100000000000000000000000000000000ffc400141001"
        "00000000000000000000000000000000ffda000c0301000210031000003f00bf80ffd9"
    )


def _xlsx_one_row(*, image_filename: str = "paneer.jpg") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "dishes"
    ws.append(BULK_HEADERS)
    ws.append(
        [
            "Bulk Paneer",
            "north_indian",
            "veg",
            199,
            30,
            20,
            50,
            "Creamy paneer",
            "Paneer, gravy",
            "Fresh daily",
            "TRUE",
            "FALSE",
            "FALSE",
            image_filename,
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_bulk_template_has_predefined_headers(client: AsyncClient, kitchen_ctx):
    _, kitchen_id, token = kitchen_ctx
    res = await client.get(
        f"/api/v1/kitchens/{kitchen_id}/dishes/bulk/template.xlsx",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]
    wb = load_workbook(io.BytesIO(res.content))
    headers = [c.value for c in next(wb["dishes"].iter_rows(min_row=1, max_row=1))]
    assert headers == BULK_HEADERS
    assert "image_filename" in headers
    assert wb["dishes"].max_row >= 3  # header + samples


@pytest.mark.asyncio
async def test_bulk_import_maps_image_filename(client: AsyncClient, kitchen_ctx):
    _, kitchen_id, token = kitchen_ctx
    xlsx = _xlsx_one_row(image_filename="paneer.jpg")
    jpeg = _minimal_jpeg()

    res = await client.post(
        f"/api/v1/kitchens/{kitchen_id}/dishes/bulk",
        headers={"Authorization": f"Bearer {token}"},
        files={
            "spreadsheet": (
                "dishes.xlsx",
                xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "images": ("paneer.jpg", jpeg, "image/jpeg"),
        },
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["accepted"] == 1
    assert body["rejected"] == 0
    assert body["images_mapped"] == 1
    assert body["results"][0]["status"] == "created"

    dish_id = body["results"][0]["dish_id"]
    assert dish_id
    # Activate is blocked without live hero — patch keeps inactive for truth-in-media.
    patched = await client.patch(
        f"/api/v1/kitchens/{kitchen_id}/dishes/{dish_id}",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "Bulk Paneer"},
    )
    assert patched.status_code == 200
    match = patched.json()
    assert match["is_active"] is False
    assert len(match["media"]) == 1
    assert match["media"][0]["is_live_capture"] is False


@pytest.mark.asyncio
async def test_bulk_import_rejects_missing_image_map(client: AsyncClient, kitchen_ctx):
    _, kitchen_id, token = kitchen_ctx
    xlsx = _xlsx_one_row(image_filename="missing.jpg")
    res = await client.post(
        f"/api/v1/kitchens/{kitchen_id}/dishes/bulk",
        headers={"Authorization": f"Bearer {token}"},
        files={
            "spreadsheet": (
                "dishes.xlsx",
                xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )
    assert res.status_code == 200
    body = res.json()
    assert body["accepted"] == 0
    assert body["rejected"] == 1
    assert "not found" in (body["results"][0]["detail"] or "").lower()


@pytest.mark.asyncio
async def test_bulk_import_accepts_images_zip(client: AsyncClient, kitchen_ctx):
    _, kitchen_id, token = kitchen_ctx
    xlsx = _xlsx_one_row(image_filename="zipped.jpg")
    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w") as zf:
        zf.writestr("zipped.jpg", _minimal_jpeg())
    res = await client.post(
        f"/api/v1/kitchens/{kitchen_id}/dishes/bulk",
        headers={"Authorization": f"Bearer {token}"},
        files={
            "spreadsheet": (
                "dishes.xlsx",
                xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "images_zip": ("photos.zip", zbuf.getvalue(), "application/zip"),
        },
    )
    assert res.status_code == 200, res.text
    assert res.json()["accepted"] == 1


def test_template_builder_unit():
    data = build_dish_bulk_template_xlsx()
    wb = load_workbook(io.BytesIO(data))
    assert "dishes" in wb.sheetnames
    assert "readme" in wb.sheetnames
