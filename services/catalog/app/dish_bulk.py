"""Owner bulk dish import — Excel template + image filename mapping (truth-in-media safe)."""

from __future__ import annotations

import io
import re
import uuid
import zipfile
from typing import Any

from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.media import store_kitchen_media_bytes
from app.models import Category, Cuisine
from app.schemas import (
    DishCreateRequest,
    DishMediaInput,
    create_dish,
    ensure_default_categories,
)
from ckac_common.event_bus import EventPublisher
from fastapi import UploadFile

MAX_BULK_ROWS = 100
MAX_BULK_IMAGES = 100

BULK_HEADERS = [
    "name",
    "cuisine_slug",
    "category_slug",
    "price",
    "prep_time_min",
    "delivery_time_min",
    "max_time_min",
    "description",
    "ingredients_description",
    "quality_measures",
    "is_featured",
    "is_chefs_special",
    "is_unique_recipe",
    "image_filename",
]

SAMPLE_ROWS = [
    {
        "name": "Paneer Butter Masala",
        "cuisine_slug": "north_indian",
        "category_slug": "veg",
        "price": 220,
        "prep_time_min": 30,
        "delivery_time_min": 20,
        "max_time_min": 50,
        "description": "Creamy tomato gravy with soft paneer.",
        "ingredients_description": "Paneer, tomato, butter, cream, spices",
        "quality_measures": "Fresh paneer daily · no artificial colour",
        "is_featured": "TRUE",
        "is_chefs_special": "FALSE",
        "is_unique_recipe": "FALSE",
        "image_filename": "paneer_butter_masala.jpg",
    },
    {
        "name": "Chicken Biryani",
        "cuisine_slug": "south_indian",
        "category_slug": "non_veg",
        "price": 280,
        "prep_time_min": 45,
        "delivery_time_min": 25,
        "max_time_min": 70,
        "description": "Slow-cooked dum biryani.",
        "ingredients_description": "Basmati rice, chicken, saffron, fried onion",
        "quality_measures": "Bone-in pieces · dum sealed",
        "is_featured": "FALSE",
        "is_chefs_special": "TRUE",
        "is_unique_recipe": "FALSE",
        "image_filename": "chicken_biryani.jpg",
    },
]


class BulkDishRowResult(BaseModel):
    row: int
    name: str | None = None
    status: str  # created | rejected
    dish_id: uuid.UUID | None = None
    detail: str | None = None


class BulkDishImportResponse(BaseModel):
    accepted: int
    rejected: int
    results: list[BulkDishRowResult] = Field(default_factory=list)
    images_mapped: int = 0
    images_unused: list[str] = Field(default_factory=list)
    note: str = (
        "Imported dishes stay inactive until you add a live-capture hero in Menu "
        "(gallery bulk photos are not live-capture)."
    )


def build_dish_bulk_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "dishes"
    ws.append(BULK_HEADERS)
    for row in SAMPLE_ROWS:
        ws.append([row.get(h, "") for h in BULK_HEADERS])

    help_ws = wb.create_sheet("readme")
    help_ws.append(["Column", "Required", "Notes"])
    help_rows = [
        ("name", "yes", "Dish name (2–255 chars)"),
        ("cuisine_slug", "yes", "e.g. north_indian, south_indian, home_style, chinese"),
        ("category_slug", "yes", "e.g. veg, non_veg, vegan, snacks, desserts, beverages"),
        ("price", "yes", "INR > 0"),
        ("prep_time_min", "no", "Default 30"),
        ("delivery_time_min", "no", "Owner-set delivery window minutes"),
        ("max_time_min", "no", "Customer-facing max; defaults to prep+delivery"),
        ("description", "no", "Plain text or simple HTML"),
        ("ingredients_description", "no", "Ingredients / allergens"),
        ("quality_measures", "no", "Hygiene / quality notes"),
        ("is_featured", "no", "TRUE/FALSE"),
        ("is_chefs_special", "no", "TRUE/FALSE"),
        ("is_unique_recipe", "no", "TRUE/FALSE"),
        (
            "image_filename",
            "recommended",
            "Exact file name of an image you upload with the sheet (e.g. paneer.jpg)",
        ),
    ]
    for r in help_rows:
        help_ws.append(list(r))
    help_ws.append([])
    help_ws.append(
        [
            "Truth in media",
            "",
            "Bulk images are draft heroes only. Dishes import inactive — "
            "open Menu and replace with a live-capture photo before activating.",
        ]
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _as_bool(value: Any, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("1", "true", "yes", "y"):
        return True
    if s in ("0", "false", "no", "n"):
        return False
    return default


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def parse_dish_bulk_xlsx(data: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if "dishes" in wb.sheetnames:
        ws = wb["dishes"]
    else:
        ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("Excel sheet is empty") from exc
    headers = [_cell_str(h).lower() for h in header_row]
    if "name" not in headers or "cuisine_slug" not in headers or "category_slug" not in headers:
        raise ValueError(
            "Excel must include columns: name, cuisine_slug, category_slug "
            "(download the sample template)"
        )
    parsed: list[dict[str, Any]] = []
    for idx, row in enumerate(rows_iter, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        record: dict[str, Any] = {"_row": idx}
        for i, key in enumerate(headers):
            if not key:
                continue
            record[key] = row[i] if i < len(row) else None
        parsed.append(record)
        if len(parsed) > MAX_BULK_ROWS:
            raise ValueError(f"Maximum {MAX_BULK_ROWS} dish rows per upload")
    return parsed


def _normalize_filename(name: str) -> str:
    base = name.replace("\\", "/").split("/")[-1].strip().lower()
    return base


async def _index_images(
    kitchen_id: uuid.UUID,
    images: list[UploadFile],
    images_zip: UploadFile | None,
) -> tuple[dict[str, str], list[str]]:
    """Upload images to media storage; return map filename→url and unused names later."""
    files: list[tuple[str, bytes, str | None]] = []

    for img in images or []:
        raw = await img.read()
        if not raw:
            continue
        fname = _normalize_filename(img.filename or f"image_{len(files)}.jpg")
        files.append((fname, raw, img.content_type))

    if images_zip and images_zip.filename:
        zdata = await images_zip.read()
        if zdata:
            with zipfile.ZipFile(io.BytesIO(zdata)) as zf:
                for info in zf.infolist():
                    if info.is_dir():
                        continue
                    name = _normalize_filename(info.filename)
                    if not re.search(r"\.(jpe?g|png|webp)$", name):
                        continue
                    files.append((name, zf.read(info), None))

    if len(files) > MAX_BULK_IMAGES:
        raise ValueError(f"Maximum {MAX_BULK_IMAGES} images per upload")

    mapping: dict[str, str] = {}
    for fname, raw, content_type in files:
        result = store_kitchen_media_bytes(
            kitchen_id=kitchen_id,
            data=raw,
            context="dish",
            is_live_capture=False,
            declared_content_type=content_type,
        )
        mapping[fname] = result.url
    return mapping, list(mapping.keys())


async def import_dishes_bulk(
    session: AsyncSession,
    kitchen_id: uuid.UUID,
    *,
    spreadsheet: UploadFile,
    images: list[UploadFile] | None,
    images_zip: UploadFile | None,
    publisher: EventPublisher | None,
) -> BulkDishImportResponse:
    xbytes = await spreadsheet.read()
    if not xbytes:
        raise ValueError("Spreadsheet file is empty")
    name_l = (spreadsheet.filename or "").lower()
    if not (name_l.endswith(".xlsx") or name_l.endswith(".xlsm")):
        raise ValueError("Upload an .xlsx Excel file (download the sample template)")

    rows = parse_dish_bulk_xlsx(xbytes)
    if not rows:
        raise ValueError("No dish rows found in the spreadsheet")

    image_map, image_keys = await _index_images(kitchen_id, images or [], images_zip)
    used_images: set[str] = set()

    await ensure_default_categories(session, kitchen_id)
    cuisines = {
        c.slug: c.id
        for c in (
            await session.execute(select(Cuisine).where(Cuisine.kitchen_id == kitchen_id))
        ).scalars().all()
    }
    categories = {
        c.slug: c.id
        for c in (
            await session.execute(select(Category).where(Category.kitchen_id == kitchen_id))
        ).scalars().all()
    }

    results: list[BulkDishRowResult] = []
    accepted = 0
    rejected = 0

    for record in rows:
        row_no = int(record.get("_row") or 0)
        dish_name = _cell_str(record.get("name"))
        try:
            cuisine_slug = _cell_str(record.get("cuisine_slug")).lower()
            category_slug = _cell_str(record.get("category_slug")).lower()
            price = _as_float(record.get("price"))
            if not dish_name or len(dish_name) < 2:
                raise ValueError("name is required")
            if price is None or price <= 0:
                raise ValueError("price must be > 0")
            cuisine_id = cuisines.get(cuisine_slug)
            if not cuisine_id:
                raise ValueError(f"Unknown cuisine_slug '{cuisine_slug}'")
            category_id = categories.get(category_slug)
            if not category_id:
                raise ValueError(f"Unknown category_slug '{category_slug}'")

            image_filename = _normalize_filename(_cell_str(record.get("image_filename")))
            media: DishMediaInput | None = None
            if image_filename:
                url = image_map.get(image_filename)
                if not url:
                    raise ValueError(
                        f"image_filename '{image_filename}' not found in uploaded images"
                    )
                used_images.add(image_filename)
                media = DishMediaInput(
                    url=url,
                    is_hero=True,
                    is_live_capture=False,
                )

            prep = _as_int(record.get("prep_time_min")) or 30
            delivery = _as_int(record.get("delivery_time_min"))
            max_time = _as_int(record.get("max_time_min"))

            # Bulk gallery photos are never live-capture → always import inactive.
            data = DishCreateRequest(
                name=dish_name,
                cuisine_id=cuisine_id,
                category_id=category_id,
                price=price,
                prep_time_min=prep,
                delivery_time_min=delivery,
                max_time_min=max_time,
                description=_cell_str(record.get("description")) or None,
                ingredients_description=_cell_str(record.get("ingredients_description")) or None,
                quality_measures=_cell_str(record.get("quality_measures")) or None,
                is_active=False,
                is_featured=_as_bool(record.get("is_featured")),
                is_chefs_special=_as_bool(record.get("is_chefs_special")),
                is_unique_recipe=_as_bool(record.get("is_unique_recipe")),
                media=media,
            )

            dish = await create_dish(session, kitchen_id, data, publisher)
            accepted += 1
            results.append(
                BulkDishRowResult(
                    row=row_no,
                    name=dish_name,
                    status="created",
                    dish_id=dish.id,
                    detail="Inactive draft — add live-capture hero before activating",
                )
            )
        except Exception as exc:
            rejected += 1
            results.append(
                BulkDishRowResult(
                    row=row_no,
                    name=dish_name or None,
                    status="rejected",
                    detail=str(exc),
                )
            )

    unused = sorted(set(image_keys) - used_images)
    return BulkDishImportResponse(
        accepted=accepted,
        rejected=rejected,
        results=results,
        images_mapped=len(used_images),
        images_unused=unused,
    )
